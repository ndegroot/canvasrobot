import re
import sys
import typing
import operator
from pathlib import Path
import logging

from pydal.objects import Row
import openpyxl
import webview
from attrs import define
import rich_click as click
from rich.progress import Progress

from result import Ok, Err, Result, is_ok, is_err  # noqa: F401
from .canvasrobot import CanvasRobot, Field
import canvasapi
from .commandline import create_db_folder, get_logger

click.rich_click.SHOW_ARGUMENTS = True
click.rich_click.MAX_WIDTH = 100
click.rich_click.TEXT_MARKUP = "rich"

MS_URL = "https://videocollege.uvt.nl/Mediasite/Play/%s"
PN_URL = "https://tilburguniversity.cloud.panopto.eu/Panopto/Pages/Viewer.aspx?id=%s"

logger = get_logger("urltransform",
                    file_level=logging.DEBUG)
logger.setLevel(logging.INFO)


def create_excel(data, file_name: str = "output.xlsx"):
    """
    Create an Excel-file for list of dict `data`

    :param data: Een list of dicts/Rows keys are fieldnames
    :param file_name: Naam van het uitvoerbestand (standaard: "output.xlsx").
    """
    if not data:
        click.echo("DEV fout: Parameter data missing")
        return

    if isinstance(data, Row):
        # make it a list
        data = [data, ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mediasite URLs"

    # column labels
    headers = list(data[0].as_dict().keys())
    ws.append(headers)

    for row in data:
        for field in row:
            if isinstance(row[field], list):
                row[field] = "\n".join(map(str, row[field]))  # some fields are list of str (or int!)
        ws.append(list(row.as_dict().values()))

    wb.save(file_name)
    click.echo(f"Excel-bestand gegenereerd: {file_name}")


def replace_and_count(original_string: str,
                      search_string: str,
                      replace_string: str,
                      dryrun=False) -> (str, int):
    count = 0
    processed_string = original_string
    while search_string in processed_string:
        processed_string = processed_string.replace(search_string, replace_string, 1)
        count += 1
    return original_string if dryrun else processed_string, count


def show_result(html: str = "", robot=None, single_course=None, dryrun=True):
    """show search result in a webview window"""

    html_with_ui = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <title>Search Results</title>
    </head>
    <body>
      <h2>Pages with Mediasite URLs</h2>
      <button onclick='pywebview.api.close()' title='Quit browser'>Close</button>
      {("<button onclick='pywebview.api.transform()' "
        "title='Perform the transformations'>Transform</button>") if dryrun else ""}
      <p>Use the link(s) to check the pages or video URLs.</p>
      <hr/>
      {html}
      <hr/>
    </body>
    </html>
    """

    class Api:
        _window = None

        def set_window(self, window):
            self._window = window

        def transform(self):
            scan_replace_urls(robot=robot,
                              single_course=single_course,
                              dryrun=False)
            self._window.html = "<p>Done</p>"

        def close(self):
            self._window.destroy()
            self._window = None

            sys.exit(0)  # needed to prevent hang
            # return count, new_body

    if not robot:
        click.echo(click.style("DEV error needs robot parameter",
                               fg="red"))
    if single_course is None:
        click.echo(click.style("DEV error needs single_course parameter",
                               fg="red"))

    api = Api()
    win = webview.create_window(title="Result report (click [Close] button to close)",
                                html=html_with_ui,
                                js_api=api)
    api.set_window(win)
    webview.start()


class ImportExcelError(Exception):
    pass


@define
class Transformation:
    """datastruct to record title, url, dryrun_state and transformed HTML of transformed
    Pages and ExternalUrls
    list contains all instances of this class
    title and url can be retrieved as a column"""

    list: typing.ClassVar[list] = []
    title: str = ""
    ctype: str = ""
    url: str = ""
    module_item_id: int = 0
    transformed: str = ""
    replacements: int = 0
    dryrun: bool = False

    def __attrs_post_init__(self):
        Transformation.list.append(self)

    @classmethod
    def get_list(cls):
        return cls.list

    @classmethod
    def clear_list(cls):
        cls.list = []

    @classmethod
    def pop(cls):
        cls.list.pop()

    @classmethod
    def get_column(cls, field):
        attr_of = operator.attrgetter(field)
        # the_list = cls.get_list()
        return list(map(attr_of, cls.list))


class UrlTransformationRobot(CanvasRobot):
    con = None
    cr = None
    current_page = None
    current_page_url = None
    # current_transformed_pages: list[TransformedPage] = []
    transformation_report = ""
    transformation_course_report = ""
    pages_changed = 0
    external_urls_changed = 0
    count_replacements = 0

    def __init__(self, db_folder: Path = None,
                 is_testing: bool = False,
                 db_auto_update: bool = False,
                 db_force_update: bool = False):
        super().__init__(db_folder=db_folder,
                         is_testing=is_testing,
                         db_auto_update=db_auto_update,
                         db_force_update=db_force_update)
        self.add_media_ids_table()
        if self.db(self.db.ids).isempty():
            self.import_ids()

    # Begin database section
    def add_media_ids_table(self):
        self.db.define_table('ids',
                             Field('panopto_id', 'string'),
                             Field('mediasite_id', 'string'))

    def import_ids(self):

        ids_result = self.get_video_ids()
        if is_err(ids_result):
            raise ImportExcelError(ids_result.err_value)

        rows = ids_result.ok_value
        print(rows[0])
        assert rows[0]['PanoptoID'] == '532f98ad-43dc-45b6-8109-aeeb01865f0e', \
            "import error in db"
        for row in rows:
            self.db.ids.insert(mediasite_id=row['MediasiteID'],
                               panopto_id=row['PanoptoID'])
        self.db.commit()

    def get_video_ids(self) -> Result[list[dict[str, str]], str]:
        """read ids table from spreadsheet"""
        xls_path = self.db_folder / "redirect_list.xlsx"
        try:
            sh = openpyxl.load_workbook(xls_path).active
            column_names = next(sh.values)[0:]
            # Initialize the list of dictionaries
            rows_as_dicts = []
            # Iterate over the sheet rows (excluding header)
            for row in sh.iter_rows(min_row=2, values_only=True):
                row_dict = {column_names[i]: row[i] for i in range(len(column_names))}
                rows_as_dicts.append(row_dict)
            return Ok(rows_as_dicts)
        except Exception as e:
            msg = f"Error opening exported video ID list({e})"
            return Err(msg)

    def lookup_panopto_id(self, mediasite_id: str) -> str:
        db = self.db  # just sugarcoat
        row = db(db.ids.mediasite_id == mediasite_id).select(db.ids.panopto_id).first()
        return row.panopto_id if row else None

    # End database section

    def mediasite2panopto(self, text: str, transformation=None, dryrun=True) -> (str, bool, int):
        """
        Replace links in a single page or other item with text
        :param transformation: info about the (possible) transformation
        :param text possibly with one or more mediasite urls
        :param dryrun: if true just statistics, no action
        :returns tuple with
        1. text with transformed mediasite urls if panopto id are found in lookup   (unless dryrun)
        2. flag True, if updates were made
        3. count of replacements made
        (replace the ms_id with p_id in the
        https://videocollege.uvt.nl/Mediasite/
        Play/ce152c1602144b80bad5a222b7d4cc731d
        replace by (redirect procedure until dec 2024)
        https://tilburguniversity.cloud.panopto.eu/Panopto/Pages/
        Viewer.aspx?id=221a5d47-84ea-44e1-b826-af52017be85c)
        """
        updated = False
        original_text = text
        count_replacements = 0
        # match each source-url and extract the id into a  list of ms_ids
        matches = re.findall(r'(https://videocollege\.uvt\.nl/Mediasite/Play/([a-z0-9]+))', text)

        if num_matches := len(matches):
            logger.debug(f"{num_matches} 'videocollege-url' matches in {self.current_page} {self.current_page_url}")
        # wrong if external_url
        msg = (f"<p><a href={transformation.url} target='_blank'>"
               f" Open {transformation.ctype} '{transformation.title}'</a></p>")
        # for each ms_id: lookup p_id and construct new target-url
        action_or_not = 'would become' if dryrun else 'changed into'
        # loop through all matches
        for match in matches:
            ms_url = match[0]
            ms_id = match[1]
            pn_id = self.lookup_panopto_id(ms_id)
            if pn_id:
                # replace source-url with target-url
                pn_url = PN_URL % pn_id

                logger.debug(f"'{ms_url}' {action_or_not} '{pn_url}' in {self.current_page_url}")

                text, count = replace_and_count(text, ms_url, pn_url, dryrun=dryrun)

                logger.debug(f"{count} occurrences {action_or_not} in '{original_text}' from {self.current_page_url}")
                msg += (f"<p><a href={ms_url} target='_blank'>{ms_url}</a>"
                        f" {action_or_not} <a href={pn_url} target='_blank'>{pn_url}</a></p>")

                count_replacements += count
                updated = True
            else:
                # no corresponding panopto id found in database
                msg += (f"<p>Page or external url "
                        f" has mediasite url {ms_url} which could NOT be transformed "
                        f"because the mediasite id is not found in DB.</p>")

                logger.warning(f"Mediasite_id {ms_id} not found {self.current_page} {self.current_page_url} {ms_id}")
            self.transformation_report += (msg + '<br/>')
            msg = "<hr/>"
            logger.debug(f"{count_replacements} candidates in {self.current_page} {self.current_page_url} ")

        return text, updated, count_replacements

    def save_transform_data_db(self, course_id: int = None):
        """
        1. using the course_id save course data in db.course
        2. save teacher data in db.course2user
        3. using the TransformedPage class save the transformed page data in db.course_urlTransform
        ( not the transformed html (or candidate when dryrun)
        """

        course = self.canvas.get_course(course_id, include=['term', 'teachers'])

        c_id = self.update_db_for(course, only_course=True)  # returns db.course.id
        # course needs to be present in course table for course2user to work

        db = self.db

        teacher_names = [teacher["display_name"] for teacher in course.teachers]
        # teacher_ids = [teacher["id"]  for teacher in course.teachers]

        # result = self.update_db_teachers(course)  # also creates db.user entries
        # if is_err(result):
        #     click.echo("No info about teachers available (authorization error)")
        #     teacher_logins, teacher_names, teacher_ids = (), (), ()
        # else:
        #     teacher_logins, teacher_names, teacher_ids = result.ok_value

        # make relational link between course-user(teacher)
        teacher_logins = list()
        teacher_emails = list()

        for teacher in course.teachers:

            user = self.get_user(teacher['id'])
            first_name, last_name, prefix = self.parse_sortable_name(user)
            try:
                profile = user.get_profile()
            except canvasapi.exceptions.Forbidden:
                logger.warning(f"Can't get profile for user {teacher['id']} in course {course_id} (Forbidden)")
                profile = dict(login_id="n.a.(due to rights)",
                               primary_email="n.a.(due to rights)")

            teacher_logins.append(profile.get("login_id"))
            teacher_emails.append(profile.get('primary_email'))
            inserted_id = db.user.update_or_insert(db.user.user_id == user.id,
                                                   user_id=user.id,
                                                   name=user.name,
                                                   first_name=first_name,
                                                   prefix=prefix,
                                                   last_name=last_name,
                                                   username=profile.get("login_id", "n.a."),
                                                   email=profile.get("primary_email", "n.a."),
                                                   role='T')
            db_user_id = inserted_id or db(db.user.user_id ==
                                           user.id).select().first().id

            _ = db.course2user.update_or_insert((db.course2user.user == db_user_id) &
                                                (db.course2user.course == c_id),
                                                user=db_user_id,
                                                course=c_id,
                                                role='T')

        list_transformations = Transformation.get_list()
        page_titles = [item.title for item in list_transformations if item.ctype == "Page"]
        module_item_ids = [item.module_item_id for item in list_transformations if item.ctype == "ExternalUrl"]

        _ = db.course_urltransform.update_or_insert(
            (db.course_urltransform.course_id == course_id),
            course_id=course_id,
            account_id=course.account_id,
            course_code=course.course_code,
            sis_code=course.sis_course_id if hasattr(course,
                                                     'sis_course_id') and course.sis_course_id else "n.a.",
            term=course.term['name'],
            name=course.name,
            teacher_logins=teacher_logins,
            teacher_names=teacher_names,
            teache_emails=teacher_emails,
            nr_pages=len(page_titles),
            nr_module_items=len(module_item_ids),
            titles=page_titles,
            urls=Transformation.get_column('url'),
            module_items=module_item_ids,
            html_report=self.transformation_course_report,
            dryrun=Transformation.dryrun
        )
        db.commit()
        pass

    # noinspection PyUnusedLocal
    def get_transform_data(self, single_course: int, all_courses=False, admin_id: int = 0) -> Row or None:
        """ get (candidate if row.dryrun) transform data as a PyDal Row
        if not single_course collect *all* available course data based on admin_id.
        if admin_id == 0 all courses for default admin-account are exported else select only course for admin_id
        :param all_courses:
        :param admin_id:
        :param single_course:
        :returns db row or None if not found"""
        # todo: admin_id
        db = self.db
        # todo: maybe optionally join with db.course/ db.user ?
        if single_course:
            row = db(db.course_urltransform.course_id == single_course).select(db.course_urltransform.ALL).first()
            return row
        if all_courses:
            # self.canvas.get_account(config.admin_id) if conf
            # self.admin = self.canvas.get_account(config.admin_id) if conf
            if admin_id == 0:
                courses = self.admin.get_courses()
            else:
                courses = self.admin.get_courses(by_subaccounts=[admin_id, ])
                if len(list(courses)) == 0:
                    # use csv file with (only) course_ids instead
                    self.console.log(f"No courses found using for admin {admin_id} access through Canvas "
                                     f"(possibly no rights). Using ids from local CSV file instead...")
                    courses = self.get_courses_admin_csv(admin_id)

            course_ids = [course.id for course in courses]

            rows = db((db.course_urltransform.course_id in course_ids) &
                      (db.course_urltransform.nr_pages > 0 or
                       db.course_urltransform.nr_module_items > 0)).select(db.course_urltransform.ALL)
            return rows

    def export_transform_data(self,
                              single_course: int = 0,
                              all_courses: bool = False,
                              admin_id: int = 0, ):
        """ export to excel """

        data = self.get_transform_data(single_course=single_course,
                                       all_courses=all_courses,
                                       admin_id=admin_id)

        create_excel(data)

    def transform_urls_in_course(self, course_id: int, dryrun=True) -> bool:
        """
        Transform the mediasite urls in all pages and module-items of the course with this course_id
        record all transformations in Transformation class-object
        :param course_id:
        :param dryrun: if true no action just candidates
        :return: True unless error
        """
        # self.transformation_course_report = ""
        logger.debug(f"Getting pages from course {course_id}")
        try:
            course, pages, module_items = self.get_course_pages_module_items(course_id)  # example

        except (Exception, canvasapi.exceptions.Forbidden) as e:
            err = f"Course {course_id} skipped due to {e}"
            logger.warning(err)
            self.errors.append(err)
            return False
        else:
            self.transformation_course_report = f"<h2>{course.id}: {course.name}</h2>"
            Transformation.clear_list()
            for page in pages:
                logger.debug(f"Handling '{page.title}'")
                if page.body:
                    transformation = Transformation(title=page.title,
                                                    url=page.html_url,
                                                    ctype="Page",
                                                    dryrun=dryrun, )
                    # builds list of page transform info
                    new_body, updated, count = self.mediasite2panopto(page.body,
                                                                      transformation=transformation,
                                                                      dryrun=dryrun)
                    self.count_replacements += count
                    if updated:
                        transformation.replacements = count
                        transformation.transformed = new_body
                        self.pages_changed += 1
                        if not dryrun:
                            # actual replacement
                            page.edit(wiki_page=dict(body=new_body))
                    else:
                        Transformation.pop()
                        # remove from Transformation.list: not an actual transformation

            ext_urls = [item for item in module_items if item.type == 'ExternalUrl']
            for ext_url in ext_urls:
                logger.debug(f"Handling '{ext_url}'")
                if ext_url.external_url:
                    transformation = Transformation(title=ext_url.title,
                                                    url=f"{self.canvas_url}/courses/"
                                                        f"{ext_url.course_id}/modules/items/{ext_url.id}",
                                                    ctype="ExternalUrl",
                                                    module_item_id=ext_url.id,
                                                    dryrun=dryrun)
                    new_url, updated, count = self.mediasite2panopto(ext_url.external_url,
                                                                     transformation=transformation,
                                                                     dryrun=dryrun)
                    if updated:
                        self.count_replacements += count
                        if not dryrun:
                            ext_url.edit(module_item=dict(external_url=new_url))

                        transformation.transformed = new_url
                        transformation.replacements = count
                        self.external_urls_changed += 1
                    else:
                        Transformation.pop()
                        # not an *actual* transformation
            self.transformation_report += ("<hr/>" + self.transformation_course_report)
            self.save_transform_data_db(course_id)
            # uses the self.transformation_course_report
            # uses the list in ClassObject Transformation added to above
        return True


@define
class TestCourse:
    """ we need an object with id attribute"""
    id: int


def go_up(path, levels=1):
    path = Path(path)
    for _ in range(levels):
        path = path.parent
    return path


@click.group(no_args_is_help=True,
             help="CLI met scan-commando met opties en een export-commando")
@click.version_option(package_name='canvasrobot')
@click.pass_context
@click.option("--db_auto_update", default=False, is_flag=True,
              help="Don't update the database automatically.")
@click.option("--db_force_update", default=False, is_flag=True,
              help="Force db update. Otherwise periodic.")
def cli(ctx: click.Context, db_auto_update, db_force_update):
    path = create_db_folder()
    robot = UrlTransformationRobot(db_auto_update=db_auto_update,
                                   db_force_update=db_force_update,
                                   db_folder=path)  # default location db: folder 'databases'
    ctx.obj = robot


@click.command(
    no_args_is_help=True,
    help="Scan for mediasite_id, no changes, replacements, report only "
         "(Unless '--just_do_it' is supplied")
@click.pass_context
@click.option("--single_course", '-s',
              default=0,
              help="Scan one course, supply Canvas id of a single course.")
@click.option("--all_courses", '-a',
              default=False,
              is_flag=True,
              help="Scan all courses.")
@click.option("--admin_id", default=0,
              help="Scan courses belonging to this admin_id (8, 20)")
@click.option("--stop_after",
              default=0,

              help="Stop scanning after this number of courses")
@click.option("--just_do_it", default=False,
              is_flag=True,
              help="Scan and [red]REPLACE[/red] the mediasite urls")
@click.version_option()
def scan(ctx, single_course, all_courses, admin_id, stop_after, just_do_it):
    dryrun = True
    if (just_do_it and
            click.confirm("Continue transforming the Mediasite urls?")):
        dryrun = False
    else:
        click.echo("Just a dryrun, no changes in Canvas.")

    robot = ctx.obj
    # not really needed
    if all_courses:
        single_course = 0

    scan_replace_urls(robot, single_course, admin_id, stop_after, dryrun)


def scan_replace_urls(robot=None,
                      single_course: int = 0,
                      admin_id: int = 0,
                      stop_after: int = 0,
                      dryrun: bool = True):
    """for a single_course (or all courses) scan and optionally replace mediasite urls
    :param robot:
    :param single_course: if 0 do all courses (for this admin_id )
    :param admin_id:
    :param stop_after:
    :param dryrun:
    """
    if single_course is None:
        click.echo(click.Style(f"DEV error '{single_course=}' should be 0 "
                               f"or a single id (not None)", fg="red"))
        return

    courses = (TestCourse(single_course),) if single_course else robot.get_courses_in_account(admin_id=admin_id)
    if stop_after:
        courses = courses[:stop_after]
    count_courses = len(courses)
    with Progress(console=robot.console) as progress:
        task_checking = progress.add_task(f"[green]checking {count_courses} courses...",
                                          total=count_courses, )
        for course in courses:
            progress.update(task_checking,
                            advance=1)  # Update progressbar

            robot.transform_urls_in_course(course.id, dryrun=dryrun)
            # updates tr.transformation_report through Transformation instances

    click.echo(f"Transformations completed ({dryrun=})")
    # conclusion
    robot.transformation_report += (f"<hr/><p>{count_courses} course{'' if count_courses == 1 else 's'} checked.</p>"
                                    f"<p>{robot.pages_changed} page(s) and {robot.external_urls_changed} external "
                                    f"url(s) "
                                    f"{'would be changed' if dryrun else 'were changed'},"
                                    f" {robot.count_replacements} urls "
                                    f"{'would be' if dryrun else ''} replaced</p>")
    if robot.transformation_report:
        report = robot.transformation_report
        show_result(report, robot, single_course, dryrun)
        # if report_html:
        with open("last_report.html", "w") as file:
            file.write(report)
    robot.report_errors()


@click.command(
    no_args_is_help=True,
    help="Export last scan results for mediasite_id in an Excelsheet")
@click.pass_context
@click.option("--single_course", '-s', default=0,
              help="Give Canvas id of a single course.")
@click.option("--all_courses", '-a',
              is_flag=True,
              default=0,
              help="Select all courses for selected admin of just all")
@click.option("--admin_id", default=0,
              help="Only export course belonging to this admin_id (8, 20)")
@click.version_option()
def export(ctx, single_course, all_courses, admin_id):
    if ('all_courses' in ctx.params.keys() and
            'admin_id' not in ctx.params.keys()):
        click.echo("Alle cursussen van alle sub-accounts worden geëxporteerd")

    robot = ctx.obj

    robot.export_transform_data(single_course=single_course,
                                all_courses=all_courses,
                                admin_id=int(admin_id))
    robot.report_errors()


cli.add_command(scan)
# cli.add_command(replace)
cli.add_command(export)

if __name__ == '__main__':
    cli()
