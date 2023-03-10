import io
import json
import mimetypes
import os
import re
import time
from collections import namedtuple
from datetime import datetime

import attrs
from attrs import define, asdict
import canvasapi
import requests
import logging

from rich.logging import RichHandler
from rich.progress import track
from canvasapi.course import Course
from canvasapi.util import combine_kwargs
from openpyxl.styles import NamedStyle, Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from .canvas_robot_model import AC_YEAR, NEXT_YEAR, ENROLLMENT_TYPES, \
    EDUCATIONS, COMMUNITIES, LocalDAL, CanvasConfig, EXAMINATION_FOLDER
from .entities import User, QuestionDTO, CourseMetadata, Grade, ExaminationDTO, Stats

logging.getLogger("canvasapi").setLevel(logging.WARNING)
# we don't need the info messages
# from this library

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler("canvasrobot.log"),
        RichHandler()
        # logging.StreamHandler(sys.stdout)
    ]
)



# noinspection PyProtectedMember
def file_update(file, **kwargs):
    """
        Updates a file.
        :calls: `PUT /api/v1/files/:id \
        <https://canvas.instructure.com/doc/api/files.html#method.files.update>`_
        :rtype: :class:`canvasapi.file.File`
        """
    response = file._requester.request(
        "PUT", "files/{}".format(file.id), _kwargs=combine_kwargs(**kwargs)
    )

    if "name" in response.json():
        response.set_attributes(response.json())

    return canvasapi.file.File(file._requester, response.json())


# define Python user-defined exceptions
class Error(Exception):
    """Base class for other exceptions"""
    pass


class DibsaRetrieveError(Error):
    """Raised when connection to Dibsa or retrieval failed"""
    pass



@define
class LabelType:
    label: str
    field_type: str


@define
class Profile:
    login_id: str
    id: int
    name: str
    short_name: str
    sortable_name: str
    avatar_url: str
    title: str
    bio: any  # can be N
    primary_email: str
    integration_id: str
    time_zone: str
    locale: str
    # effective_locale str)
    # calendar str)
    # lti_user_id str)



# noinspection PyTypeChecker,PyUnresolvedReferences,PyCallingNonCallable
class CanvasRobot(object):
    db: callable
    TOT_WEIGHT: int = 100
    year: int = AC_YEAR
    def __init__(self, reset_api_keys=False, years_back = 0, msg_queue=None):
        self.year = AC_YEAR - years_back
        self.queue = msg_queue
        config = CanvasConfig(reset_api_keys=reset_api_keys)
        self.canvas = canvasapi.Canvas(config.url, config.api_key)
        self.admin = self.canvas.get_account(config.admin_id) if config.admin_id \
            else None
        db_path = os.path.join(os.getcwd(),'databases')
        if not os.path.exists(db_path):
            # Create a new directory because it does not exist
            os.makedirs(db_path)
        self.db = LocalDAL()
        self.teacher_ids = self.lookup_teachers_db()
        self.internal_id = None
        logging.info("Canvasrobot started")
        self.errors = []
        self.actions = []

    def add_to_queue(self, msg, value):
        if self.queue:
            self.queue.put((msg,value))
    def lookup_teachers_db(self):
        db = self.db
        teachers = db((db.course2user.role == 'T') &
                      (db.course2user.user == db.user.id)).select(db.user.user_id,
                                                                  distinct=True)
        teacher_ids = [teacher.user_id for teacher in teachers]
        return teacher_ids

    # COURSES----------------------------------------
    def get_course(self, course_id: int):
        """"":returns canvas course by its id"""
        return self.canvas.get_course(course_id)

    def get_courses(self, enrollment_type: str = "teacher"):
        """"
        :enrollment_type 'teacher'(default), 'student','ta', 'observer' 'designer'
        :returns canvas courses for current user in role"""
        return self.canvas.get_courses(enrollment_type=enrollment_type)

    def get_courses_in_account(self, by_teachers: list=None, this_year=True):
        """
        get all course in account here use_is has the role/type [enrollment_type]
        :param by_teachers: list of teacher id's
        :param this_year: True=filter courses to include only the current year
        :returns list of courses
        """
        assert self.admin, "No (sub) admin account provided"
        courses = []
        for course in self.admin.get_courses(by_teachers=by_teachers):
            # only show/insert/update course if current year

            if this_year and (str(course.sis_course_id)[:4] != str(self.year)
                              or course.name.endswith('conclude')):
                continue
            courses.append(course)
        return courses

    # USER ----------------------------------------
    def get_user(self, user_id: int):
        """
        get user using
        :param user_id:
        :returns user
        """
        return self.canvas.get_user(user_id)

    @staticmethod
    def create_profile(profile_dict):
        try:
            profile = Profile(**profile_dict)
        except TypeError as e:
            logging.error(f"attribute needs to be added to Profile class: {e}")
            raise
        return profile

    def get_user_profile_id(self, user_id: int):
        user = self.get_user(user_id)
        profile = self.create_profile(user.get_profile())
        return user, profile

    def get_user_profile_anr(self, anr: str):
        user = self.canvas.get_user(anr, 'sis_user_id')
        profile = self.create_profile(user.get_profile())
        return user, profile

    def is_teacher_canvas(self, user):
        """":param user """
        role_teacher = ENROLLMENT_TYPES['teacher']
        # check if param: user is teacher in one of the TST courses
        for course in self.tst.get_courses():
            enrollments = course.get_enrollments()
            for enrollment in enrollments:
                if enrollment.role == role_teacher and enrollment.user == user:
                    return True
        return False

    def is_teacher_db(self, user):
        """":param user"""
        return user.id in self.teacher_ids

    def show_modules(self, course_id):
        course = self.get_course(course_id)
        modules = course.get_modules()
        # print(dir(modules))
        for page in modules:
            print(dir(page))
            print("page:{}".format(page))
            module_items = page.get_module_items()
            print("page module_items:{}".format(module_items))
            for item in module_items:
                print(item.c)
            # print("There are {} modules in page".format(len(page)))
        return

    def course_metadata(self, course_id, examination_names=None, show_candidates=False):
        """
        return dict with metadata of this course
        as a dict.
        :returns md: CourseMetadata"""
        examination_names = examination_names or []
        course = self.get_course(course_id)
        modules = course.get_modules()

        nr_modules = len(list(modules))
        nr_module_items = 0
        nr_ext_urls = 0
        for module in modules:
            # print(dir(module))
            # print("page:{}".format(module))
            module_items = module.get_module_items()
            # print("page module_items:{}".format(module_items))
            # for item in module_items:
            #    print(item.type)
            item_count_ext_url = filter(lambda i: i.type == "ext_url", module_items)
            nr_ext_urls += len(list(item_count_ext_url))
            nr_module_items += module.items_count
        pages = course.get_pages()
        pages = filter(lambda p: p.title[0:3] != 'UVT', pages)
        nr_pages = len(list(pages))
        assignments = course.get_assignments()
        assignments_summary = "Assignments:\n"
        examination_candidates = []
        for assignment in assignments:
            if assignment.name in examination_names:
                assignments_summary += f"Examination assignment: {assignment.name}"
            else:
                assignments_summary += f"Examination candidate: {assignment.name}" \
                    if show_candidates else ""

                candidate = ExaminationDTO(course_id,
                                           course.name,
                                           assignment.name)

                examination_candidates.append(candidate)
            submissions_summary = ""
            submissions = assignment.get_submissions()
            try:
                for idx, submission in enumerate(submissions, start=1):
                    if assignment.name in examination_names:
                        if submission.submission_type == "online_upload":

                            originality_str = (f"{submission.has_originality_report}"
                            if hasattr(submission, 'has_originality_report')
                            else "no has_originality_report attribute!\n")
                            submissions_summary += (f"({idx}. "
                                                    f"{submission.submission_type}) "
                                                    f"graded "
                                                    f"{submission.grade} at "
                                                    f"{submission.graded_at}. "
                                                    f"Checked for plagiarism: "
                                                    f"{originality_str}"
                                                    )
                        else:
                            submissions_summary+=(f"({idx}. "
                                                  f"{submission.submission_type}) "
                                                  f"graded "
                                                  f"{submission.grade} at "
                                                  f"{submission.graded_at}\n")
            except BaseException as exc:
                logging.exception(f"In Course:{course_id} {course.name} "
                                  f"Assignment {assignment.name} "
                                  f"Submission nr {idx} "
                                  f"type{submission.type} {exc}")
                raise


            assignments_summary+=f"\n{submissions_summary}"
        nr_assignments = len(list(assignments))
        quizzes = course.get_quizzes()
        nr_quizzes = len(list(quizzes))
        files = course.get_files()
        nr_files = len(list(files))
        # check for uploaded examination files
        examinations_summary = ""
        examination_files = 0
        for file in files:
            folder = course.get_folder(file.folder_id)
            if f"/{EXAMINATION_FOLDER}" in folder.full_name:
                examination_files+=1
                examinations_summary+=f"\n{file.display_name}"
        if examination_files == 0:
            examinations_summary += (f"\nNo examination files in "
                                     f"folder {EXAMINATION_FOLDER}")
        else:
            examinations_summary += (f"\nTotal: {examination_files} "
                                     f"examination files")
        # was this working earlier 2.2.0 ?
        # try:
        #    collaborations = course.get_collaborations()
        #    list_of_cols = [c for c in collaborations]
        #    nr_collaborations = len(list_of_collaborations)
        # except TypeError:
        #    nr_collaborations = None
        md = CourseMetadata(nr_modules=nr_modules,
                            nr_module_items=nr_module_items,
                            nr_pages=nr_pages,
                            nr_assignments=nr_assignments,
                            nr_quizzes=nr_quizzes,
                            nr_files=nr_files,
                            assignments_summary=assignments_summary,
                            examinations_summary=examinations_summary,
                            examination_candidates=examination_candidates
                            )
        # examination_candidates: collect to create a canonical list
        return md

    def enroll_in_course(self,
                         search: str,
                         course_id: int,
                         username: str,
                         enrollment_type: str) -> str:
        if search:
            try:
                course = self.get_course_using_osiris_id(search)
            except canvasapi.exceptions.ResourceDoesNotExist:
                return f"Course {search} not found in Canvas"
        else:
            course = self.get_course(course_id)

        user = self.search_user(username)
        if not user:
            # search_user / get_user fails to find a newly imported account
            # work around it by using enroll_user
            try:
                course.enroll_user(f"sis_login_id:{username}", enrollment_type)
            except Exception as e:
                return f"Failed with {e} while using special syntax for sys_login_id"
            return f"Enrolled {user} in {course} as {enrollment_type}"

        try:
            course.enroll_user(user, enrollment_type)
        except Exception as e:
            return f"Enroll of {user} in {course} as {enrollment_type} failed:{e}"
        return f"Enrolled {user} in {course} as {enrollment_type}"

    def get_all_active_tst_courses(self, from_db=True):
        """"":returns list of all TST course canvas course """

        def cur_year_active(course):
            """ filter function"""
            return (str(course.sis_course_id)[:4] == str(self.year) and
                    not course.name.endswith('conclude'))

        def set_id_to_course_id(course):
            course.id = course.course_id
            return course

        if from_db:
            db = self.db
            courses = db(
                (db.course.ac_year == self.year)).select(db.course.ALL)
            # map(set_id_to_course_id, courses)
        else:
            courses = self.tst.get_courses()
            courses = filter(cur_year_active, courses)
        # for index, c in enumerate(courses):
        #     print(index, c.name)
        return courses

    def get_course_using_osiris_id(self, osiris_id):
        """
        :returns first TST course with sisid starting with
        osiris_id in current ac. year"""
        for course in self.tst.get_courses():
            # only consider course if current year
            if str(course.sis_course_id)[:4] != str(self.year):
                continue
            if course.course_code.startswith(osiris_id):
                return course

    def get_course_id_using_osiris_id_from_db(self, osiris_id: str):
        db = self.db
        rows = db((db.course.course_code == osiris_id) &
                  (db.course.ac_year == self.year)).select(db.course.course_id)
        if rows:
            return rows[0].course_id
        return 0

    def search_user(self, search_name: str, email: str = None):
        """
        :param search_name:
        :param email to search on email
        try search name as a login, then the email if supplied otherwise use
        """
        try:
            user = self.canvas.get_user(search_name, 'sis_login_id')
        except (canvasapi.exceptions.Unauthorized,
                canvasapi.exceptions.ResourceDoesNotExist):
            if not email:
                # out of options
                self.errors.append(f'Unable to lookup {search_name} '
                                   f'using email, parameter not provided')
                return False
            try:
                user = self.canvas.get_user(email, 'email')
            except canvasapi.exceptions.ResourceDoesNotExist:
                self.errors.append(f'Unable to lookup {search_name} using email')
                return False  # give up
            except canvasapi.exceptions.Unauthorized:
                self.errors.append(f"User {search_name} "
                                   f"not (alllowed to be) found by email in Canvas")
                return False  # give up
        return user

    @staticmethod
    def get_students_dibsa(c_name):
        """get students from (local) DIBSA CRM"""
        # idea:  use LDAP instead
        url = f'http://127.0.0.1:8000/dibsa/service/call/json/students/{c_name}'
        try:
            r = requests.get(url)
        except requests.exceptions.ConnectionError:
            raise DibsaRetrieveError(f"Unable to connect to source "
                                     f"Dibsa @ {url} maybe first start web2py locally")
        try:
            result = r.json()
        except json.decoder.JSONDecodeError:
            raise DibsaRetrieveError(f"unable to decode student data "
                                     f"of {c_name} from Dibsa @ {url}")
        except Exception as e:
            raise DibsaRetrieveError(f"unable to decode student data "
                                     f"of {c_name} from Dibsa @ {url} due to {e}")

        return result

    def get_students_for_community(self, c_id):

        community_edu_ids = {'banl': ['banl', 'pm-ma', 'pm-ulo'],  # nl
                             'bauk': ['bauk', 'pm-macs'],  # uk
                             'ma': ['ma'],
                             'ulo': ['ulo'],
                             'macs': ['macs'],
                             'acskills': [edu.lower() for edu in EDUCATIONS]
                             }
        # edu_ids = [edu.lower() for edu in EDUCATIONS]
        # if c_id == 'acskills' else  if c_id == 'banl' else [c_id]
        edu_ids = community_edu_ids[c_id]
        students_dibsa = []
        students_canvas = []
        for edu_id in edu_ids:
            try:
                students_dibsa += self.get_students_dibsa(edu_id.upper())
            except DibsaRetrieveError as e:
                logging.error(e)
                raise
        for student in students_dibsa:
            username = student['username']
            user = self.search_user(username)
            if not user:
                continue
            students_canvas.append(user)

        return students_canvas

    def get_community(self, c_id):
        # c_id can be an education (each education has a community)
        # or acskills (all educations)
        # Decide: add pm-ma to BA?
        #         pm-ulo to ULO?
        #         pm-macs to MACS
        try:
            course = self.get_course(COMMUNITIES[c_id])
        except IndexError:
            return None, ["wrong c_id"]

        return course

    def enroll_user_in_course(self, user, course, role):
        try:
            course.enroll_user(user, role)
        except (canvasapi.exceptions.BadRequest,
                canvasapi.exceptions.Conflict):
            self.errors.append(f'User {user.name} not added to {course.name}')
        else:
            self.actions.append(f"{user.name} added to {course.name} as {role}")
        return

    def add_observer_to_education(self, user, edu_id, report_only=False):
        """ add user as an observer to all courses of an education"""
        # todo: select courses using membership of education using osiris ids
        print(user)
        # idea: filter course of an education using db
        for course in self.tst.get_courses():
            # only insert/update course if current year
            if str(course.sis_course_id)[:4] != str(self.year):
                continue
            observers = course.get_users(enrollment_type="observer")
            for observer in observers:
                if not hasattr(observer, 'login_id'):
                    continue
                print(observer.name)

    def remove_observer_from_all_courses(self, username):
        """ remove user as an observer from all TST courses"""
        removed = []
        try:
            user = self.canvas.get_user(username, 'sis_login_id')
        except canvasapi.exceptions.ResourceDoesNotExist:
            return f"User {username} not found in Canvas"

        for course in self.tst.get_courses():
            # only get a course if current year
            if str(course.sis_course_id)[:4] != str(self.year):
                continue
            enrollments = course.get_enrollments(type="ObserverEnrollment")
            # specifying user_id is not allowed
            for enrollment in enrollments:
                if enrollment.user_id != user.id:
                    continue
                print(enrollment)
                enrollment.deactivate(task='delete')
                removed.append(course.name)

        return removed

    # noinspection PyCallingNonCallable
    def import_courses(self, filename):
        """from csv file updates table Course and Teacher
        :param filename: fname csv file
        """

        import csv
        from os.path import expanduser

        path = os.path.join(expanduser('~'), 'Downloads', filename + '.csv')
        db = self.db
        try:
            ofile = open(path, "rU")
        except IOError as e:
            logging.error("{} {} not found".format(e.message, path))
            raise
        else:
            reader = csv.reader(ofile, delimiter=';', quotechar='"')
            header = reader.next()
            logging.debug(header)
            # db(db.teacher.id > 0).delete()
            db(db.course2teacher.id > 0).delete()
            # refresh all course2teacher couplings
            for row in reader:
                logging.debug(row)
                assert len(row) == 7, "Error {0} fields!".format(len(row))
                teacher_names = row[2].replace(' (T)',
                                               '').replace(' (U)',
                                                           '').replace(' en ',
                                                                       ', ').split(', ')
                # phases = row[4].split(', ')
                course_id = db.course.update_or_insert(db.course.course_base == row[0],
                                                       course_base=row[0],
                                                       name=row[1],
                                                       teacher_names=row[2],
                                                       ects=int(row[3]),
                                                       phase=row[4],
                                                       department=row[5],
                                                       memo=row[6])
                if not course_id:  # must be there
                    course_id = db(db.course.course_base ==
                                   row[0]).select(db.course.id)[0].id
                else:
                    logging.info("course added")
                for t_name in teacher_names:
                    teacher_id = db.teacher.update_or_insert(db.teacher.name == t_name,
                                                             name=t_name)
                    if not teacher_id:
                        teacher_id = db(db.teacher.name ==
                                        t_name).select(db.teacher.id)[0].id
                    db.course2teacher.insert(course=course_id,
                                             teacher=teacher_id)

            db.commit()
            ofile.close()
        return

    # ----------------------------------------
    # Other interactions

    def execute_command(self, command, params=None):
        """
        :param command: command to execute
        :param params: parameters to tune command
        """
        self.open()  # open canvas object
        try:
            url = self.base_url + (self.commands[command].format(params)
                                   if params else self.commands[command])
        except Exception:
            raise NotImplementedError(f"error in command {command} or params {params}")
        else:
            try:
                self.browser.get(url)
            except Exception as e:
                raise ValueError("self.browser.get uses url {} e:{}".format(url, e))
            return self.browser

    # the commands ------------------------------------------------------------

    def ud_course_users(self):
        total_count = 0
        for bb_id in self.get_course_bb_ids():
            count, count_students = self.ud_students(bb_id=bb_id)
            total_count += count

        return total_count

    def ud_students(self, bb_id=None):
        """ add enrolled users for given or currently selected course to db
        :param bb_id: if not given use self.internal_id
        """

        db = self.db
        if bb_id:
            # set self,internal_id using lookup
            self.internal_id = bb_id
        else:
            assert self.internal_id
        # we need the primary key
        try:
            bbcourse_id = db(db.bbcourse.bb_id ==
                             self.internal_id).select(db.bbcourse.id).first().id
        except AttributeError:
            logging.info(f'course_id {self.internal_id} not found in our database!')
            return -1

        logging.info(f'updating users for course_id {bbcourse_id}')
        # show overview enrolled users
        self.execute_command('enrolled_users', self.internal_id)
        # assumes course is selected
        count = 0
        count_students = 0
        # connect user to this course in db
        db.commit()

        return count, count_students

    def get_courses_data(self):
        """ for all courses: get a matrix and labels coursename and other fields from db
        :return: features-matrix, labels """

        db = self.db  # cosmetic reasons
        qry = db.course.ac_year == self.year
        # couluns = ()
        courses = db(qry).select(db.course.course_id,
                                 db.course.nr_modules,
                                 db.course.nr_module_items,
                                 db.course.nr_pages,
                                 db.course.nr_assignments,
                                 orderby=db.course.id)
        labels = [c.course_id for c in courses]
        features = [(c.nr_modules,
                     c.nr_module_items,
                     c.nr_pages,
                     c.nr_assignments) for c in courses]
        columns = ("nrModules", "nrModuleItems", ",nrPAges", "nr_assignments")
        return features, labels, columns

    def get_bbcourses(self, single_course=None):
        """ for all courses: get coursename and other fields from db
        :param single_course: used for testing
        :return: rows/list of dicts """

        db = self.db  # cosmetic reasons
        suffix = '-{}-'.format(self.year)
        qry = db.bbcourse.course_suffix.contains(suffix)
        if single_course:
            qry &= db.bbcourse.course_base == single_course
        rows = db(qry).select(db.bbcourse.bb_id,
                              db.bbcourse.name,
                              orderby=db.bbcourse.name)
        return rows

    def report_studentcounts(self):
        """ for all courses: get coursename and student count
        :return: list of dicts, keys fname, count"""
        # join bbcourses with users
        db = self.db  # cosmetic reasons
        count = db.bbuser.id.count()
        rows = db((db.course2user.bbcourse_id == db.bbcourse.id) &
                  (db.course2user.bbuser_id == db.bbuser.id) &
                  (db.course2user.role == 'S')).select(db.bbcourse.name,
                                                       count,
                                                       groupby=db.bbcourse.name,
                                                       orderby=db.bbcourse.name)
        for row in rows:
            row.count = row[count]
        return rows

    def update_all_file_urls(self, max_courses=None, single_course=None):
        # type: (int, string) -> string
        """ for all courses harvest the attachments/files urls from Bb
        :param single_course: for testing
        :param max_courses:
        :returns list of found files
        """
        courses = self.get_bbcourses(single_course=single_course)
        logging.info("{} courses to scan".format(len(courses)))

        total_files = []
        for count, course in enumerate(courses):
            if max_courses and count == max_courses:
                logging.info("Stopped after {} courses".format(max_courses))
                break
            logging.info("{}/{}:{}".format(count + 1, len(courses), course.name))
            self.goto(course.bb_id)
            areas = self.get_areas()  # top level areas (menu buttons)
            logging.info("#{} areas#".format(len(areas)))
            for area in areas:
                self.goto_area(area)  # select content area
                bb_files = self.get_files_from_area(level=0, area_name=area.name)
                # above function is recursive
                logging.info("{} files#".format(len(bb_files)))
                # test_file_name = bb_files[0].fname
                self.update_documents(bb_files)  # record data files in database
                total_files += bb_files
        return "{} courses scanned {} files found".format(len(courses),
                                                          len(total_files))

    # the delegates ---
    # noinspection PyRedeclaration
    def make_enroll_file(self, courseid):
        self.db.make_enroll_file(courseid)

    def _check_bb(self, suffix=NEXT_YEAR):
        self.db.check_bb(suffix)

    def show_user(self, user_id):
        self.execute_command('showuser', user_id)
        time.sleep(120)

    def update_documents(self, files):
        """ insert file object properties (fname, url) in table bbdocument
        :param files: list of file objects ( containing fname, url)
        """
        db = self.db
        for c_file in files:
            # lookup bbcourse id
            bbcourse_id = db(db.bbcourse.bb_id ==
                             self.internal_id).select(db.bbcourse.id).first().id
            try:
                if c_file.level > 1:
                    db.bbdocument.update_or_insert(db.bbdocument.url == c_file.url,
                                                   url=c_file.url,
                                                   name=c_file.name,
                                                   bbcourse_id=bbcourse_id,
                                                   area=c_file.area_name,
                                                   check_status=0)
                else:
                    db.bbdocument.update_or_insert(db.bbdocument.url == c_file.url,
                                                   url=c_file.url,
                                                   name=c_file.name,
                                                   bbcourse_id=bbcourse_id,
                                                   area=c_file.area_name)
            except Exception as e:
                logging.error(
                    "'{0}' error in update_documents() while inserting "
                    "'{1}' in table Document".format(e, c_file.name))
            else:
                db.commit()

    def get_list_of_documents(self):
        """ get documents/attachments from db als a list
        """
        db = self.db
        suffix = '-{}-'.format(self.year)
        try:
            items = db((db.bbcourse.course_suffix.contains(suffix)) &
                       (db.bbdocument.bbcourse_id ==
                        db.bbcourse.id)).select(db.bbdocument.ALL)
        except Exception as e:
            logging.error("*{0} error in get_list_of_ documents() "
                          "while selecting documents*".format(e))
            return []
        else:
            return items

    # noinspection PyUnresolvedReferences
    def download_documents_localfs(self):
        db = self.db
        counters = namedtuple('Counters', ['total', 'ok', 'failed'])
        items = self.get_list_of_documents()
        counters.total = len(items)
        counters.ok = 0
        counters.failed = 0
        for item in items:
            status = self.download_file(fname=item.name, url=item.url)
            if status == 200:
                counters.ok += 1
            else:
                counters.failed += 1
            db(db.bbdocument.id == item.id).update(http_status=status)
        db.commit()
        return counters

    # noinspection PyUnresolvedReferences
    def transfer_files_to_server(self):
        """transfer all documents in the bbdocument table to the server"""
        db = self.db
        counters = namedtuple('Counters', ['total', 'ok', 'failed'])
        items = self.get_list_of_documents()
        counters.total = len(items)
        counters.ok = 0
        counters.failed = 0
        for item in items:
            status = self.transfer_file_to_server(fname=item.name, url=item.url)
            if status == 200:
                counters.ok += 1
            else:
                counters.failed += 1
            db(db.bbdocument.id == item.id).update(http_status=status)
        db.commit()
        return counters

    def get_examinations_from_database(self, candidate=False):
        db = self.db
        # include the NULL values
        qry = ((db.examination.id>0)&
               (db.examination.candidate==
                candidate)) if candidate else (db.examination.id > 0)
        records = db( qry ).select(db.examination.ALL)
        return records
    def get_courses_from_database(self,
                                  skip_courses_without_students=False,
                                  qry=None,
                                  orderby=None,
                                  fields=None):
        db = self.db
        if skip_courses_without_students:
            cur_qry = (db.course.nr_students>0)
        else:
            cur_qry = (db.course.id>0)

        if qry:
            cur_qry = cur_qry & qry

        fields = fields or db.course.ALL
        orderby=orderby or db.course.course_code
        records = db(cur_qry).select(db.course.ALL,
                                     orderby=orderby)

        return records
    def get_course_from_database(self,course_id):
        db = self.db
        record = db(db.course.course_id==
                    course_id).select(db.course.ALL).first()
        return record
    def delete_course_from_database(self,course_id):
        db = self.db
        result = db(db.course.course_id==course_id).delete()
        db.commit()
        return result

    def update_record_db(self, search_field, search_id, table ,field, value):
        db = self.db
        ud_fields = {field: value}
        #row = db(db[table][search_field] == search_id).select()
        #row2 = db(db.course.course_id == search_id).select()
        db(db[table][search_field] == search_id).update(**ud_fields)
        db.commit()

    # noinspection PyUnusedLocal
    def update_database_from_canvas(self,
                                    single_course=None,
                                    max_number=None,
                                    stop_list=None):
        """
            Using the canvasapi to read the list of TST courses and
            - record internal_id course_id, fname and instructors in the
            table course
            - put teacher details in table user
            - collects info about assignments in assignment_summary
            - collects info about examinations in examination_summary
            - reports new tentamination candidates
            :return number of added/updated rows
            """

        db = self.db
        msg= f'open courselist for year {self.year} - {self.year+1}'
        self.add_to_queue(msg, None)
        logging.info(msg)
        num_rows = 0
        # tst = self.canvas.get_account(6)  # admin account
        courses=[self.get_course(single_course),] if single_course \
            else self.admin.get_courses()
        num_courses = len(list(courses))
        max_number = max_number or num_courses

        for idx,course in enumerate(courses):
            self.add_to_queue("<Progress>", (course.name,idx,num_courses))

            # only insert/update course if current year unless single_course
            if (str(course.sis_course_id)[:4] != str(self.year)
                    or course.name.endswith('conclude')) and not single_course:
                continue
            if course.name in (stop_list or []):
                continue
            if idx > max_number:
                break
            logging.debug("course: {}".format(course.name))  # course
            students = course.get_users(enrollment_type="student")
            nr_students=len(list(students))
            teachers = course.get_users(enrollment_type="teacher")
            teachers_ids = []
            for teacher in teachers:
                if not hasattr(teacher, 'login_id'):
                    continue
                # print(teacher.name)
                first_name, last_name, prefix = self.parse_sortable_name(teacher)
                inserted_id = db.user.update_or_insert(db.user.username ==
                                                       teacher.login_id,
                                                       user_id=teacher.id,
                                                       name=teacher.name,
                                                       first_name=first_name,
                                                       prefix=prefix,
                                                       last_name=last_name,
                                                       username=teacher.login_id,
                                                       email=teacher.email,
                                                       role='T')
                teachers_ids.append(inserted_id or
                                    db(db.user.username ==
                                       teacher.login_id).select().first().id)
            try:
                # skips teachers with non-accepted invites
                # (they don't have a login attribute)
                teacher_logins = [teacher.login_id for teacher in teachers
                                  if hasattr(teacher, 'login_id')]
                teacher_names = [teacher.name for teacher in teachers
                                 if hasattr(teacher, 'login_id')]
            except AttributeError as e:
                msg = (f"skipped teacher of {course.name} "
                       f"[{course.id}] due to {e}")
                logging.warning(msg)
                continue
            logging.info("instructors: {0}".format(teacher_logins))  # instructors

            inserted_id = None
            format_str = "%Y-%m-%dT%H:%M:%SZ"
            creation_date = datetime.strptime(course.created_at, format_str)
            course_id = None
            examinations = self.get_examinations_from_database(candidate=False)
            canonical_examination_names = [row.name for row in examinations
                                           if (row.course==course.id and not row.candidate)]
            md = self.course_metadata(course.id, canonical_examination_names)
            try:
                course_id = db.course.update_or_insert(db.course.course_id == course.id,
                                                       course_id=course.id,
                                                       # status=course.workflow_state,
                                                       course_code=course.course_code.split('-')[0],
                                                       sis_code=course.sis_course_id,
                                                       # year course_code and suffixes
                                                       name=course.name,
                                                       creation_date=creation_date,
                                                       ac_year=self.year,
                                                       nr_students=nr_students,
                                                       nr_modules=md.nr_modules,
                                                       nr_module_items=md.nr_module_items,
                                                       nr_pages=md.nr_pages,
                                                       nr_assignments=md.nr_assignments,
                                                       nr_quizzes=md.nr_quizzes,
                                                       assignments_summary= md.assignments_summary,
                                                       examinations_summary = md.examinations_summary,
                                                       teachers=teacher_logins,
                                                       teachers_names=teacher_names)
            except Exception as e:
                err = f"{e} error inserting {course.name}"
                self.add_to_queue("<InsertError>", err)
                logging.exception(err)
                raise
            if course_id:
                db(db.course.id == inserted_id).update(status=2)
            else:
                course_id = db(db.course.course_id == course.id).select().first().id
            for cand in md.examination_candidates:
                # candidate is True if course_name in examination_list else False
                db.examination.update_or_insert((db.examination.course ==
                                                 cand.course_id)&
                                                 (db.examination.name == cand.name),
                                                 course=cand.course_id,
                                                 course_name=cand.course_name,
                                                 name=cand.name,
                                                )
            for user_id in teachers_ids:
                db.course2user.update_or_insert((db.course2user.course == course_id) &
                                                (db.course2user.user == user_id),
                                                course=course_id,
                                                user=user_id,
                                                role='T')

            db.commit()
            num_rows += 1

        self.add_to_queue("<Done>",
                          (f"Update db from Canvas "
                           f"{single_course or max_number or 'All courses'}"))
        return num_rows

    def is_user_valid(self, userinfo):
        """"":param userinfo (dict or named tuple or Storage with attributes id, login_id)
             :returns True for invalid user, detail"""
        if type(userinfo) == dict:
            # import collections
            # User = collections.namedtuple('User', 'id')
            user = User(id=userinfo['id'])
        else:
            user = userinfo

        user = self.canvas.get_user(user.id)
        profile = user.get_profile()
        if 'login_id' not in profile.keys():
            return True, 'no login_id, assume valid'
        if 'closed' in profile['login_id']:
            return False, 'login closed'
        if 'invalid' in profile['primary_email']:
            return False, 'mail invalid'
        return True, 'appears valid'

    @staticmethod
    def valid_c_id(c_id):
        valid_c_ids = [edu for edu in EDUCATIONS]
        valid_c_ids.append('ACSKILLS')
        if c_id.upper() not in valid_c_ids:
            return False, [f"Wrong c_id '{c_id.upper()}' should "
                           f"be one of {valid_c_ids} in lower case"]
        return True, ""

    @staticmethod
    def deactivate_enrollment(enrollment):
        enrollment.deactivate('delete')

    def deactivate_if_invalid(self, enrollment, report_only):
        valid, reason = self.is_user_valid(enrollment.user)
        if valid:
            return "no action"
        user_name = enrollment.user['fname']
        if report_only:
            self.actions.append(f"{user_name} would be deactivated ({reason})")
            return "would be deactivated"
        try:
            self.deactivate_enrollment(enrollment)
        except Exception as e:
            self.errors.append(f"Deactivate of {user_name} failed:{e}")
            return "deactivation failed"
        else:
            self.actions.append(f"{user_name} deactivated ({reason})")
        return "deactivated"

    def get_enrollments(self, c_id):
        course = self.get_course(COMMUNITIES[c_id])
        enrollments = course.get_enrollments()

        # errors = []

        def filter_student(enrollment):
            """"filter the real students"""
            try:
                result = (enrollment.user['name'].lower() != 'test student' and
                          enrollment.role == 'StudentEnrollment')
            except Exception as e:
                self.errors.append(f"{enrollment.user['name']} not retrieved {e}")
                return False
            return result

        # enrollments = [enrollment for enrollment in enrollments]
        enrollments = filter(filter_student, enrollments)
        return enrollments

    def cleanup_community(self, c_id, report_only):
        result = self.check_c_cid(c_id)
        if result != 'ok':
            return result

        course = self.get_course(COMMUNITIES[c_id])
        enrollments = course.get_enrollments()
        errors = []
        removed = []
        for enrollment in enrollments:
            if enrollment.user['fname'].lower() == 'test student':
                continue
            if enrollment.role == 'StudentEnrollment':
                try:
                    invalid, reason = self.user_invalid(enrollment.user)
                except Exception as e:
                    errors.append(f"{enrollment.user['fname']} not retrieved {e}")
                    continue
                if invalid:

                    if not report_only:
                        try:
                            enrollment.deactivate('delete')
                        except Exception as e:
                            errors.append(e)
                            continue
                    removed.append((enrollment.user['fname'], reason))

        num_removed = len(removed)
        msg = (f"{num_removed} students would have been removed"
               if report_only
               else f"{num_removed} students removed")

        return removed, errors, msg

    @staticmethod
    def parse_sortable_name(user):
        """examples:
        Klein, Wim
        Groot, Nico de
        Goyvaert, Samuel (Sam)
        Wieringen, Archibald (H.M.J.)
        """

        assert ", " in user.sortable_name, "sortable_name should contain comma"
        source = user.sortable_name
        pat = re.compile(
            r'(?P<last_name>[\w \-]+), (?P<first_name>\w+)\s?((\((?P<first_name_par>\w+)\))|'
            r'(\((?P<init>[\w.]+.)\)))?(\s*(?P<prefix>\w+))?')
        d = re.match(pat, source)
        if not d:
            print("!!!Failed parsing!!!")

        return d['first_name_par'] or d['first_name'], d['prefix'] or '', d['last_name']

    def download_file(self, url=None, fname=None):
        """download a file to static folder from url
        :param fname: local file to write to
        :param url: remote url to fetch file from
        """
        filename = 'static/' + fname.strip()
        logging.info("creating file %s for %s" % (filename, url))
        # convert to fname/value pairs
        cookies = {}
        for cookie in self.cookies:
            cookies[str(cookie['fname'])] = cookie['value']
        with open(filename, 'wb') as outfile:
            r = requests.get(url, cookies=cookies, stream=True)
            if r.status_code == 200:
                for block in r.iter_content(1024):
                    if not block:
                        break
                    outfile.write(block)
            else:
                logging.info("html error %s" % r.status_code)
        return r.status_code

    def transfer_file_to_server(self, url=None, fname=None):
        """download a file to uploads folder from external url using a temporary buffer
        :param fname: original name of file
        :param url: remote url to fetch file from
        """
        db = self.db
        filename = fname.strip()
        logging.debug("creating transfer buffer %s for %s" % (filename, url))
        # convert to fname/value pairs
        cookies = {}
        for cookie in self.cookies:
            cookies[str(cookie['fname'])] = cookie['value']
        #
        logging.info("cookies %s for %s" % (cookies, url))
        with io.BytesIO() as outfile:
            try:
                r = requests.get(url, cookies=cookies, stream=True)
            except requests.exceptions.RequestException as e:
                logging.error("couldn't get {} from file {} due to {}".format(filename,
                                                                              url,
                                                                              e.message))
                return None

            if r.status_code == 200:
                self.receive_file(db, filename, outfile, r, url)
            else:
                logging.info("html response {} url is now {}".format(r.status_code, r.url))
                # try again with redirected url (we assume redirection)
                r = requests.get(r.url, cookies=cookies, stream=True)
                if r.status_code == 200:
                    self.receive_file(db, filename, outfile, r, url)
                else:
                    logging.error('Failed {0}'.format(r))

        return r.status_code

    @staticmethod
    def receive_file(db, filename, outfile, r, url):
        logging.debug('Start receiving file in buffer')
        for block in r.iter_content(1024):
            if not block:
                logging.debug('.')
                break
            outfile.write(block)
        outfile.seek(0)  # rewind
        content_type = r.headers['content-type']
        full_filename = 'undef'
        try:
            full_filename = filename + mimetypes.guess_extension(content_type) or ''
            thisfile = db.bbdocument.bbfile.store(outfile, full_filename)
        except TypeError as e:
            msg = "unable to store {} with content-type {} {}".format(filename,
                                                                      content_type,
                                                                      e.message)
            logging.error(msg)
        except IOError as e:
            msg = "unable to store {} based on {} and {} {}".format(full_filename,
                                                                    filename,
                                                                    content_type,
                                                                    e.message)
            logging.error(msg)
        else:
            logging.info("{} should be in uploads folder...".format(full_filename))
            db(db.bbdocument.url == url).update(bbfile=thisfile)
            db.commit()

    def course_grades(self, c_id):
        course = self.get_course(c_id)
        enrollments = course.get_enrollments()
        grades = []
        for enrollment in enrollments:
            if enrollment.user['fname'].lower() == 'test student':
                continue
            if enrollment.role == 'StudentEnrollment':
                # print(enrollment)
                grades.append(Grade(stud_name=enrollment.user['fname'],
                                    stud_id=enrollment.user['login_id'],
                                    final_score=enrollment.grades['final_score'],
                                    final_grade=enrollment.grades['final_grade']))

        return course, grades

    def get_grades(self, c_id):
        """" From the course
        :param c_id
        get the grades and create an Excel file for Osiris import"""
        course, grades = self.course_grades(c_id)

        labels_and_types = [LabelType(label='stud_id', field_type='string'),
                            LabelType(label='stud_name', field_type='string'),
                            LabelType(label='score', field_type='percentage'),
                            LabelType(label='grade', field_type='grade')]
        styles = {
            # list below is not complete
            # see https://www.web2pyref.com/reference/field-type-database-field-types
            'datetime': NamedStyle(name='date', number_format='yyyy-mm-dd'),
            'id': NamedStyle(name='int', number_format='#,##0'),
            'integer': NamedStyle(name='int', number_format='#,##0'),
            'price': NamedStyle(name='money', font=Font(italic=True),
                                fill=PatternFill(fill_type='solid', fgColor='404040'),
                                number_format=u'[$\u20ac-1] #,##0.00 '),  # unicode for ???
            'double': NamedStyle(name='double', number_format='#0.000000'),
            'grade': NamedStyle(name='grade', number_format='#0.00'),
            'percentage': NamedStyle(name='score', number_format='#0.00'),
            'string': NamedStyle(name='std',
                                 alignment=Alignment(horizontal='left',
                                                     vertical='bottom',
                                                     text_rotation=0,
                                                     wrap_text=False,
                                                     shrink_to_fit=False,
                                                     indent=0)),
            'text': None
        }

        dest_filename = f'TST {course}.xlsx'
        wb = Workbook()
        ws = wb.active
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToHeight = 0
        ws.page_setup.fitToWidth = 1
        # add 6 describing rows
        ws.append(["Faculteit", "", "TST"])
        ws.append(["Vakbenamimg", "", course.name])
        ws.append(["Vakcode", "", course.course_code])
        ws.append(["Datum tentamen", "", "[in te vullen]"])
        ws.append([])
        ws.append(["Student", "SIS User ID", "Final Score", "Final Grade"])
        # data
        for row in grades:
            ws.append(attrs.astuple(row))

        # set width
        dim_holder = DimensionHolder(worksheet=ws)
        widths = [25, 12, 12, 12]
        for index, col in enumerate(range(ws.min_column, ws.max_column + 1)):
            dim_holder[get_column_letter(col)] = ColumnDimension(ws,
                                                                 min=col,
                                                                 max=col,
                                                                 width=widths[index])
        ws.column_dimensions = dim_holder

        # apply the styles, must be cell by cel
        for row in ws.iter_rows(min_row=2):
            for index, cell in enumerate(row):
                try:
                    lookup = labels_and_types[index].field_type  # lookup
                    cell.style = styles[lookup]  # lookup
                except ValueError:
                    pass

        # mark first six row as header
        for cell in ws[6]:
            cell.style = 'Pandas'
        ws.freeze_panes = "A7"
        ws.print_title_rows = '1:6'
        wb.save(filename=dest_filename)

        return f"Excel file {dest_filename} created with {len(grades)} rows", dest_filename

    def create_quiz(self, course_id: int, title: str, quiz_type: str = '') -> int:
        """
        :param course_id:
        :param title:
        :param quiz_type:
        :return: msg, quiz_id
        """
        course: Course = self.get_course(course_id)
        quiz = course.create_quiz(dict(title=title, quiz_type=quiz_type))
        logging.debug(f"{course.name} now contains {quiz}")
        return quiz.id

    def create_question(self,
                        course_id: int,
                        quiz_id: int,
                        question_dto: QuestionDTO) -> int:
        """
        :param course_id:
        :param quiz_id:
        :param question_dto:
        :return: msg, quiz_question_id
        """
        course = self.get_course(course_id)
        quiz = course.get_quiz(quiz_id)
        quiz_question = quiz.create_question(question=asdict(question_dto))
        logging.debug(f"{quiz} now contains {quiz_question}")
        return quiz_question.id
        # return f"{quiz} now contains {quiz_question}", quiz_question.id

    def create_quizzes_from_data(self,
                                 course_id: int,
                                 question_format="Question {}.",
                                 data=None,
                                 gui_root=None,
                                 gui_queue=None
                                 ):
        """
        :param course_id: Canvas course_id: the quizzes are added to this course
        :param question_format: used to create the question name.
        They will be numbered. Should contain '{}' as  placeholder
        starting with 1
        :param data: the quizdata
        :param gui_root: used in combination with GUI (tkinter)
        :param gui_queue: used in combination with GUI
        :return: stats:Stats
        """
        if '{}' not in question_format:
            msg = (f"parameter 'question_format(={question_format})' "
                   f"should contain {{}} als placeholder")
            ValueError(msg)

        stats = Stats()

        total_questions = 0
        for _, questions in data:
            total_questions+=len(questions)

        for quiz_name, questions in track(data): # for non-gui progress

            quiz_id = self.create_quiz(course_id=course_id,
                                       title=quiz_name,
                                       quiz_type="practice_quiz")
            stats.quiz_ids.append(quiz_id)
            for index, (question_text, answers) in enumerate(questions, start=1):
                # answers_asdict = [asdict(answer) for answer in answers]
                question_dto = QuestionDTO(question_name=question_format.format(index),
                                           question_text=question_text,
                                           answers=answers)
                question_id = self.create_question(course_id=course_id,
                                                   quiz_id=quiz_id,
                                                   question_dto=question_dto)
                stats.question_ids.append(question_id)
            if gui_root:
                gui_queue.put(index/total_questions)
                gui_root.event_generate('<<CreateQuizzes:Progress>>')

        if gui_root:
            gui_root.event_generate('<<CreateQuizzes:Done>>')

        return stats


    def get_course_tab_by_label(self, course_id: int, label: str):
        course = self.get_course(course_id)
        for tab in course.get_tabs():
            if tab.label == label:
                return tab

    def create_folder_in_course_files(self, course_id: int, foldername: str):
        course = self.get_course(course_id)

        foldernames = [f.full_name for f in course.get_folders()]
        if f"course files/{foldername}" in foldernames:
            logging.info(f"No action needed: Folder '(course) files/{foldername}' "
                         f"already in ({course_id})")
            return -1
        folder_id = course.create_folder(foldername,
                                         parent_folder_path='/',
                                         locked=True)
        logging.info(f"Folder '(course) files/{foldername}' should be created now")
        return folder_id

    def create_folder_in_all_courses(self, foldername):
        for course in track(self.get_all_active_tst_courses(from_db=False),
                            description="All current courses..."):
            self.create_folder_in_course_files(course.id, foldername)

    def unpublish_subfolder_in_all_courses(self,
                                           foldername: str,
                                           files_too: bool = False,
                                           check_only: bool = False):
        for course in track(self.get_all_active_tst_courses(from_db=False),
                            description=(f"Checking all current"
                                         f" courses for folder '{foldername}'..." if check_only
                            else f"Unpublish all published folder {foldername}...")):
            if course.name.endswith("_conclude"):
                continue
            # logging.info(course.name)
            self.unpublish_folderitems_in_course(course.id,
                                                             foldername,
                                                             files_too,
                                                             check_only)

    def unpublish_folderitems_in_course(self, course_id: int,
                                        foldername: str,
                                        files_too: bool = False,
                                        check_only: bool = False):
        """
        :param check_only: only show publication status no change
        :param foldername:
        :param course_id:
        :param files_too: if true unpublish files in the folder too (recursive)
        :returns  list of course ids of courses with missing foldername
        """
        file_changes = 0
        folder_changes = 0

        def unpublish_items(file_or_folder):
            nonlocal folder_changes
            nonlocal file_changes
            nonlocal course_id
            for folder in file_or_folder.get_folders():
                if not folder.locked:
                    folder.update(locked=True)
                    folder_changes += 1
                    logging.info(f"Corrected: Folder '{folder.full_name}' "
                                 f"is now unpublished!")
                unpublish_items(folder)
            for file in file_or_folder.get_files():
                if not file.locked:
                    if not check_only:
                        file_update(file, locked=True)
                        logging.info(f"Corrected: File '{file.display_name}' in {foldername} "
                                     f"is now unpublished")
                        file_changes += 1
                    else:
                        logging.warning(f"File '{file.display_name}' in {foldername} is published!")

        # files_folder = 'course files'
        course_ids_missing_folder = []
        course = self.get_course(course_id)
        folders = course.get_folders()
        for folder in folders:  # paginated
            if folder.full_name == foldername:
                files_tab = self.get_course_tab_by_label(course_id, "Files") or \
                            self.get_course_tab_by_label(course_id, "Bestanden")
                try:
                    if files_tab.visibility == "public":
                        logging.warning(f"Files folder of {course.name}"
                                        f" ({course.id}) is visible")
                        # files_tab.visibility = "admins" # ! no change  without teachers approval!
                except AttributeError:
                    logging.warning(f"Files tab visibility of {course.name} {course_id} missing")

                # folder_id = folder.id
                if not folder.locked:
                    if not check_only:
                        folder.update(locked=True)
                        folder_changes += 1
                        logging.info(
                            f"Folder '{foldername}' is now unpublished"
                            f" in course {course.name}")
                    else:
                        logging.warning(f"Folder '{foldername}' is published "
                                        f"in course {course.name}({course.id})!")
                else:
                    logging.debug(
                        f"Folder '{foldername}' was already "
                        f"unpublished/locked in course {course.name}")
                if files_too:
                    unpublish_items(folder)
                    # for f in folder.get_files():
                    #    if not f.locked:
                    #        logging.warning(f"{f.filename} is published!")
                break
        else:
            logging.info(f"Folder '{foldername}' not found in {course.name}")
            course_ids_missing_folder.append(course_id)

        for course_id in course_ids_missing_folder:
            logging.info(f"Folder '{foldername}' not found in {course_id}")
        return folder_changes, file_changes


if __name__ == '__main__':
    pass
